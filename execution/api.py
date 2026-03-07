import time
import io
import logging
import asyncio
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from fastapi import FastAPI, HTTPException, File, Form, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from dotenv import load_dotenv

# Uvicorn runs this file from the project root, so imports must be absolute starting from the 'execution' module.
from execution.scrape_instagram import scrape_profile
from execution.generate_titles import generate_titles
from execution.generate_cover import generate_cover_zip
from execution.manage_leads import (
    create_lead, get_lead, get_lead_by_number, get_all_leads,
    update_status, update_lead, append_conversation,
)

load_dotenv()

# --- Logging configuration ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# --- File upload constraints ---
MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}

app = FastAPI(title="Instagram Profile Extractor")

# Mount static files directory
app.mount("/static", StaticFiles(directory="static"), name="static")

class URLsRequest(BaseModel):
    urls: List[str]

class ProspectRequest(BaseModel):
    instagram_url: str
    whatsapp_number: str
    cover_path: str = ""
    formatted_name: str = ""
    specialty_line: str = ""
    headline: str = ""
    username: str = ""

class ManualMessageRequest(BaseModel):
    message: str

progress_store = {}

@app.get("/progress")
def progress_api(task_id: str):
    return progress_store.get(task_id, {"current": 0, "total": 0})

@app.get("/", response_class=FileResponse)
def serve_frontend():
    return FileResponse("static/index.html")

@app.post("/extract")
async def extract_profiles(request: URLsRequest):
    logger.info("Extract request received for %d URLs", len(request.urls))
    results = []

    for i, url in enumerate(request.urls):
        # 2-second delay between requests as per directive, except for the first one
        if i > 0:
            await asyncio.sleep(2)

        profile_data = await asyncio.to_thread(scrape_profile, url)

        if "error" in profile_data:
            # If error scraping, append error and continue
            logger.warning("Scrape error for '%s': %s", url, profile_data["error"])
            results.append({
                "url": url,
                "error": profile_data["error"],
                "error_type": profile_data.get("error_type", "unknown")
            })
            continue

        # GPT call for titles
        gpt_data = await asyncio.to_thread(generate_titles, profile_data["name"], profile_data["bio"])

        results.append({
            "username": profile_data["username"],
            "name": profile_data["name"],
            "bio": profile_data["bio"],
            "external_link": profile_data["external_link"],
            "formatted_name": gpt_data.get("formatted_name", ""),
            "specialty_line": gpt_data.get("specialty_line", ""),
            "headline": gpt_data.get("headline", "")
        })

    logger.info("Extract completed: %d results", len(results))
    return results

@app.post("/extract-batch")
async def extract_batch(request: URLsRequest, task_id: str = ""):
    logger.info("Batch extract request received for %d URLs (task_id=%s)", len(request.urls), task_id or "none")
    if task_id:
        progress_store[task_id] = {"current": 0, "total": len(request.urls)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Username", "Nome", "Especialidade", "Headline", "Link Externo"])

    for i, url in enumerate(request.urls):
        if i > 0:
            await asyncio.sleep(2)

        profile_data = await asyncio.to_thread(scrape_profile, url)

        if "error" in profile_data:
            ws.append([profile_data.get("username", url), "ERRO", profile_data["error"], "", ""])
        else:
            gpt_data = await asyncio.to_thread(generate_titles, profile_data["name"], profile_data["bio"])
            ws.append([
                profile_data["username"],
                gpt_data.get("formatted_name", ""),
                gpt_data.get("specialty_line", ""),
                gpt_data.get("headline", ""),
                profile_data["external_link"]
            ])

        if task_id:
            progress_store[task_id]["current"] = i + 1
            
    # Format header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
            
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Optional cleanup to avoid memory leak for progress dictionary
    if task_id in progress_store:
        # Give JS a moment to read 100%
        # Ideally, a background task cleans this up, but for this utility keeping it in memory is negligible.
        pass
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultados.xlsx"}
    )

@app.post("/generate-cover")
async def generate_cover(
    photo: UploadFile = File(...),
    brand_color: str = Form(...),
    instagram_url: str = Form(...)
):
    # --- File upload validation ---
    if photo.content_type and photo.content_type not in ALLOWED_IMAGE_TYPES:
        logger.warning("Rejected upload: invalid content type '%s'", photo.content_type)
        raise HTTPException(status_code=400, detail=f"Invalid file type: {photo.content_type}. Allowed: JPEG, PNG, WebP.")

    photo_bytes = await photo.read()

    if len(photo_bytes) > MAX_UPLOAD_SIZE:
        size_mb = len(photo_bytes) / (1024 * 1024)
        logger.warning("Rejected upload: file too large (%.1f MB)", size_mb)
        raise HTTPException(status_code=400, detail=f"File too large ({size_mb:.1f} MB). Maximum allowed: 10 MB.")

    if len(photo_bytes) < 100:
        raise HTTPException(status_code=400, detail="File appears to be empty or corrupted.")

    try:
        logger.info("Generating cover for '%s'", instagram_url)
        zip_bytes = await asyncio.to_thread(generate_cover_zip, photo_bytes, brand_color, instagram_url)
        return StreamingResponse(
            io.BytesIO(zip_bytes),
            media_type="application/zip",
            headers={"Content-Disposition": 'attachment; filename="capa.zip"'}
        )
    except Exception as e:
        logger.error("Cover generation failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

# --- POST /prospect --- creates a lead in the database ---
@app.post("/prospect")
async def prospect_lead(request: ProspectRequest):
    """
    Creates a new lead with status = cover_generated.
    Called after /generate-cover to register the lead for prospecting.
    """
    logger.info("Prospect request for '%s' (%s)", request.instagram_url, request.whatsapp_number)

    # Check for duplicate number
    existing = get_lead_by_number(request.whatsapp_number)
    if existing:
        logger.warning("Duplicate prospect: number '%s' already exists (lead_id=%d)",
                       request.whatsapp_number, existing["id"])
        raise HTTPException(
            status_code=409,
            detail="Lead com este número de WhatsApp já existe (ID: {}).".format(existing["id"])
        )

    try:
        lead = create_lead(
            instagram_url=request.instagram_url,
            username=request.username,
            whatsapp_number=request.whatsapp_number,
            formatted_name=request.formatted_name,
            specialty_line=request.specialty_line,
            headline=request.headline,
            cover_path=request.cover_path,
        )
        logger.info("Lead created via /prospect: id=%d", lead["id"])
        return {"status": "queued", "lead_id": lead["id"]}
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        logger.error("Error creating lead: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

# --- POST /webhook/whatsapp --- receives inbound messages from Z-API ---
@app.post("/webhook/whatsapp")
async def webhook_whatsapp(payload: dict):
    """
    Z-API webhook receiver.
    Branch A (warm_up_sent): classify -> personalized reply -> sequence
    Branch B (message_sent+): conversation mode agent -> reply -> status eval
    """
    # Z-API sends phone as 'phone' and message text as 'text.message' or 'text'
    phone = payload.get("phone", "")
    # Handle different Z-API payload structures
    text_data = payload.get("text", {})
    if isinstance(text_data, dict):
        message = text_data.get("message", "")
    else:
        message = str(text_data)

    if not phone or not message:
        logger.debug("Webhook received non-text or empty payload, ignoring")
        return {"status": "ignored"}

    logger.info("Webhook received from %s: '%s'", phone, message[:100])

    # Find the lead by phone number
    lead = get_lead_by_number(phone)
    if lead is None:
        logger.info("Webhook from unknown number %s, ignoring", phone)
        return {"status": "unknown_number"}

    lead_id = lead["id"]
    current_status = lead["status"]

    # Always append the inbound message to conversation history
    append_conversation(lead_id, "lead", message)
    update_lead(lead_id, last_lead_reply_at=datetime.now(timezone.utc).isoformat())

    # --- Branch A: Lead just responded to warm-up ---
    if current_status == "warm_up_sent":
        logger.info("Branch A: Lead %d responded to warm-up", lead_id)
        update_status(lead_id, "warm_up_responded")

        # Classification + sequence will be handled by conversational_agent.py (P7)
        # For now, log the event. Once P7 is implemented, this will:
        # 1. Call classification_mode to detect contact type (A/B/C/D)
        # 2. Save contact_classification to DB
        # 3. Send personalized reply
        # 4. Dispatch 5-part sequence via send_whatsapp.send_sequence()
        try:
            from execution.conversational_agent import handle_warm_up_response
            await asyncio.to_thread(handle_warm_up_response, lead_id, message)
        except ImportError:
            logger.warning("conversational_agent.py not yet available — Branch A queued for lead %d", lead_id)
        except Exception as e:
            logger.error("Error in Branch A for lead %d: %s", lead_id, e)

        return {"status": "branch_a_processed", "lead_id": lead_id}

    # --- Branch B: Lead replied during/after sequence ---
    elif current_status in ("message_sent", "awaiting_response", "in_conversation"):
        logger.info("Branch B: Lead %d replied (status=%s)", lead_id, current_status)

        # Transition to in_conversation if not already there
        if current_status in ("message_sent", "awaiting_response"):
            if current_status == "message_sent":
                update_status(lead_id, "awaiting_response")
            update_status(lead_id, "in_conversation")

        # Conversation mode will be handled by conversational_agent.py (P7)
        try:
            from execution.conversational_agent import handle_conversation_reply
            await asyncio.to_thread(handle_conversation_reply, lead_id, message)
        except ImportError:
            logger.warning("conversational_agent.py not yet available — Branch B queued for lead %d", lead_id)
        except Exception as e:
            logger.error("Error in Branch B for lead %d: %s", lead_id, e)

        return {"status": "branch_b_processed", "lead_id": lead_id}

    # --- Terminal or unhandled status ---
    else:
        logger.info("Webhook for lead %d in status '%s', no action taken", lead_id, current_status)
        return {"status": "no_action", "lead_id": lead_id, "current_status": current_status}


# --- GET /leads --- list all leads for dashboard ---
@app.get("/leads")
async def list_leads():
    """Returns all leads for the P10 dashboard."""
    leads = get_all_leads()
    # Mask numbers slightly or just return them if it's an internal dashboard
    return {"leads": leads}


# --- POST /leads/{id}/message --- manual takeover sending ---
@app.post("/leads/{lead_id}/message")
async def send_manual_message(lead_id: int, request: ManualMessageRequest):
    """Allows a human to send a message to the lead via the dashboard (P11)."""
    lead = get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
        
    try:
        from execution.send_whatsapp import send_followup
        success = send_followup(lead_id, request.message, role="agent (manual)")
        if success:
            return {"status": "success"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send message via Z-API.")
    except Exception as e:
        logger.error("Error sending manual message for lead %d: %s", lead_id, e)
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
